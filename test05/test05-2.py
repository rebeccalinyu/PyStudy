from openpyxl import load_workbook
wb = load_workbook(filename='sample.xlsx', read_only=True)
ws = wb['Testdata']

wb1 = load_workbook(filename='sample-1.xlsx')
ws1 = wb1['Testdata-1']

ws1.cell(row=7, column=7, value=ws['A1'].value)		

wb1.save("sample-1.xlsx")
