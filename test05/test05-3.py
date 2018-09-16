from openpyxl import load_workbook
wb = load_workbook(filename='samplenew.xlsx', read_only=True)
ws = wb['OS Exp']

wb1 = load_workbook(filename='samplenew-1.xlsx')
ws1 = wb1['OS&Travel Exp']

ws1.cell(2, 25, value=ws['H3'].value)		

wb1.save("samplenew-1.xlsx")
